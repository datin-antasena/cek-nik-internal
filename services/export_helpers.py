import io

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _write_info_sheet(writer, info_rows):
    if not info_rows:
        return

    df_info = pd.DataFrame(info_rows)
    df_info.to_excel(writer, index=False, sheet_name="INFO_PROSES")
    info_ws = writer.book["INFO_PROSES"]
    info_ws.column_dimensions["A"].width = 28
    info_ws.column_dimensions["B"].width = 100
    for col_letter in ("A", "B"):
        for cell in info_ws[col_letter]:
            cell.number_format = "@"


def buat_excel_buffer(df_result, selected_sheet, info_rows=None):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_name = f"Cek_{selected_sheet}"[:30]
        df_result.to_excel(writer, index=False, sheet_name=sheet_name)

        ws = writer.book[sheet_name]
        for idx in range(len(df_result.columns)):
            col_letter = get_column_letter(idx + 1)
            ws.column_dimensions[col_letter].width = 25
            for cell in ws[col_letter]:
                cell.number_format = "@"
        _write_info_sheet(writer, info_rows)

    buffer.seek(0)
    return buffer


def buat_merge_excel_buffer(df_result, info_rows=None):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_result.to_excel(writer, index=False, sheet_name="MERGED_DATA")

        ws = writer.book["MERGED_DATA"]
        for idx in range(len(df_result.columns)):
            col_letter = get_column_letter(idx + 1)
            ws.column_dimensions[col_letter].width = 25
            for cell in ws[col_letter]:
                cell.number_format = "@"

        _write_info_sheet(writer, info_rows)

    buffer.seek(0)
    return buffer


def _format_text_sheet(writer, sheet_name, width=25):
    ws = writer.book[sheet_name]
    for idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width
        for cell in ws[col_letter]:
            cell.number_format = "@"


def buat_merge_error_report_buffer(error_frames: dict[str, pd.DataFrame]):
    buffer = io.BytesIO()
    sheet_map = {
        "REKAP_ERROR": error_frames.get("error_summary_df", pd.DataFrame()),
        "DUPLIKAT": error_frames.get("duplicate_df", pd.DataFrame()),
        "KUNCI_KOSONG": error_frames.get("empty_key_df", pd.DataFrame()),
        "KOLOM_WAJIB_KOSONG": error_frames.get("required_empty_df", pd.DataFrame()),
    }

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheet_map.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            _format_text_sheet(writer, sheet_name, width=28 if sheet_name == "REKAP_ERROR" else 25)

    buffer.seek(0)
    return buffer


def buat_validation_error_report_buffer(error_frames: dict[str, pd.DataFrame]):
    buffer = io.BytesIO()
    sheet_map = {
        "REKAP_ERROR": error_frames.get("summary_df", pd.DataFrame()),
        "DUPLIKAT": error_frames.get("duplicate_df", pd.DataFrame()),
        "SUDAH_SALUR": error_frames.get("salur_df", pd.DataFrame()),
        "KOSONG": error_frames.get("empty_df", pd.DataFrame()),
        "TIDAK_VALID": error_frames.get("invalid_df", pd.DataFrame()),
        "USIA_TIDAK_VALID": error_frames.get("usia_invalid_df", pd.DataFrame()),
    }

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheet_map.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            _format_text_sheet(writer, sheet_name, width=28 if sheet_name == "REKAP_ERROR" else 25)

    buffer.seek(0)
    return buffer


def bersihkan_nama_file(nama):
    for ext in (".xlsx", ".xlsm", ".xls", ".csv"):
        if nama.endswith(ext):
            return nama[: -len(ext)]
    return nama


def sanitize_excel_sheet_name(name, existing_names=None):
    invalid_chars = set('[]:*?/\\')
    cleaned = "".join("-" if ch in invalid_chars else ch for ch in str(name).strip())
    cleaned = cleaned.strip("'")
    cleaned = cleaned or "Sheet"
    cleaned = cleaned[:31]

    existing = set(existing_names or [])
    if cleaned not in existing:
        return cleaned

    base = cleaned[:28] if len(cleaned) > 28 else cleaned
    counter = 2
    candidate = f"{base}_{counter}"
    while candidate in existing:
        counter += 1
        suffix = f"_{counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
    return candidate


def _enforce_text_format_in_memory(excel_bytes: bytes, sheet_name: str, selected_columns: set) -> bytes:
    return _enforce_text_format_for_sheets_in_memory(
        excel_bytes,
        {sheet_name: selected_columns},
    )


def _enforce_text_format_for_sheets_in_memory(excel_bytes: bytes, sheet_columns_map: dict[str, set]) -> bytes:
    wb = load_workbook(io.BytesIO(excel_bytes))

    for sheet_name, selected_columns in sheet_columns_map.items():
        if sheet_name not in wb.sheetnames or not selected_columns:
            continue

        ws = wb[sheet_name]
        col_indices = {}
        for col_cell in ws[1]:
            if col_cell.value in selected_columns:
                col_indices[col_cell.column] = col_cell.value

        if not col_indices:
            continue

        for col_idx in col_indices:
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        cell.value = str(cell.value).strip()
                        cell.number_format = "@"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

