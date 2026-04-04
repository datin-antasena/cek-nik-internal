import io
import os
import re as _re
import zipfile
import time
from datetime import datetime, date, timedelta as _timedelta
from difflib import get_close_matches as _get_close_matches
from zoneinfo import ZoneInfo

from dateutil import parser as dateutil_parser

import gspread
import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook
from google.oauth2.service_account import Credentials


# ─── CONFIG ───────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Dashboard Validasi Data Sentra Antasena", layout="wide")

STYLES = """
<style>
.footer {
    position: fixed;
    left: 0; bottom: 0;
    width: 100%;
    background-color: #f8f9fa;
    color: #6c757d;
    text-align: center;
    padding: 10px;
    font-size: 13px;
    border-top: 1px solid #dee2e6;
    z-index: 1000;
}
.stApp { margin-bottom: 80px; }

[data-testid="stMetricValue"] {
    font-size: 2rem;
    font-weight: bold;
    color: #0d6efd;
}
.stCheckbox {
    background-color: #e2e3e5;
    padding: 10px;
    border-radius: 5px;
    border: 1px solid #ced4da;
}
.stCheckbox label p,
.stCheckbox label span {
    color: #000000 !important;
    font-weight: bold;
}
</style>
<div class="footer">
    Dikembangkan oleh <strong>POKJA DATA DAN INFORMASI</strong>
    untuk digunakan internal <strong>Antasena</strong>
</div>
"""

COLOR_MAP = {
    "UNIK":               "#28a745",
    "GANDA":              "#dc3545",
    "BUKAN ANGKA":        "#ffc107",
    "TIDAK 16 DIGIT":     "#fd7e14",
    "TERKONVERSI (000)":  "#17a2b8",
    "KOSONG":             "#6c757d",
    "SUDAH SALUR 2026":   "#6f42c1",
}

COLOR_MAP_KATEGORI = {
    "ANAK":   "#4fc3f7",
    "DEWASA": "#66bb6a",
    "LANSIA": "#ffa726",
    "TIDAK VALID": "#ef5350",
}

TEXT_COLUMNS_KEYWORDS = [
    "NIK", "KK", "NO KK", "NO. HP", "NOMOR HP", 
    "TANGGAL", "TGL", "SK", "NOMOR SK", 
    "BAST", "NOMOR BAST", "NOMOR", "NO "
]


# ─── DATA FETCHING ─────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600)
def ambil_data_salur_gspread():
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
        creds = Credentials.from_service_account_info(
            st.secrets["gcp_service_account"], scopes=scopes
        )
        client = gspread.authorize(creds)
        sheet = client.open_by_key(st.secrets["SPREADSHEET_ID"]).worksheet("BNBA")

        kolom_nik = sheet.col_values(4)
        set_nik_salur = {str(nik).strip() for nik in kolom_nik[1:] if nik}
        waktu_update = datetime.now(ZoneInfo("Asia/Jakarta")).strftime("%d %b %Y, %H:%M:%S WIB")

        return set_nik_salur, waktu_update

    except Exception as e:
        return set(), f"Gagal mengambil data: {e}"


# ─── LOGGING ──────────────────────────────────────────────────────────────────

def catat_log(nama_file, nama_sheet, rincian_per_kolom):
    waktu = datetime.now(ZoneInfo("Asia/Jakarta")).strftime("%Y-%m-%d %H:%M:%S")
    summary_text = ""

    for col, stats in rincian_per_kolom.items():
        simple_stats = {}
        ganda_total = 0
        for k, v in stats.items():
            if str(k).startswith("GANDA"):
                ganda_total += v
            else:
                simple_stats[k] = v
        if ganda_total > 0:
            simple_stats["GANDA (TOTAL)"] = ganda_total

        stat_str = ", ".join(f"{k}:{v}" for k, v in simple_stats.items())
        summary_text += f"[{col}: {stat_str}] "

    pesan = (
        f"[{waktu}] FILE: {nama_file} | SHEET: {nama_sheet} | DETAIL: {summary_text}\n"
    )
    with open("activity_log.txt", "a") as f:
        f.write(pesan)


# ─── VALIDATION LOGIC ─────────────────────────────────────────────────────────

def cek_validitas(row, col_name, temp_col, referensi_salur):
    val = str(row[col_name]).replace(".0", "").strip()
    count = row[temp_col]

    if not val:                         return "KOSONG"
    if len(val) != 16:                  return "TIDAK 16 DIGIT"
    if not val.isdigit():               return "BUKAN ANGKA"
    if val.endswith("000"):             return "TERKONVERSI (000)"
    if val in referensi_salur:          return "SUDAH SALUR 2026"
    if count == 1:                      return "UNIK"
    return f"GANDA {count}"


def proses_kolom(df_result, col_name, use_auto_clean, referensi_salur):
    df_result[col_name] = df_result[col_name].replace("nan", "")

    if use_auto_clean:
        df_result[col_name] = (
            df_result[col_name]
            .str.replace(r"\.0$", "", regex=True)
            .str.replace(r"\D", "", regex=True)
        )
    else:
        df_result[col_name] = df_result[col_name].str.strip()

    temp_col = f"__temp_count_{col_name}"
    df_result[temp_col] = df_result.groupby(col_name).cumcount() + 1

    ref = referensi_salur if "NIK" in col_name.upper() else set()
    status_col = f"STATUS_{col_name}"
    df_result[status_col] = df_result.apply(
        lambda row: cek_validitas(row, col_name, temp_col, ref), axis=1
    )
    df_result.drop(columns=[temp_col], inplace=True)

    return df_result


# ─── AGE CATEGORY LOGIC ───────────────────────────────────────────────────────

_BULAN_ID = {
    "januari": "January", "februari": "February", "maret": "March",
    "april": "April", "mei": "May", "juni": "June",
    "juli": "July", "agustus": "August", "september": "September",
    "oktober": "October", "november": "November", "desember": "December",
}

def _ganti_bulan_id(tgl_str: str) -> str:
    kata_list = _re.findall(r"[a-zA-Z]{3,}", tgl_str)
    hasil = tgl_str
    for kata in kata_list:
        kata_lower = kata.lower()
        if kata_lower in _BULAN_ID:
            hasil = _re.sub(kata, _BULAN_ID[kata_lower], hasil, flags=_re.IGNORECASE)
        else:
            cutoff = 0.75 if len(kata_lower) <= 5 else 0.6
            cocok = _get_close_matches(kata_lower, _BULAN_ID.keys(), n=1, cutoff=cutoff)
            if cocok:
                hasil = _re.sub(kata, _BULAN_ID[cocok[0]], hasil, flags=_re.IGNORECASE)
    return hasil

_FORMATS_PASTI = [
    "%Y/%m/%d", "%Y-%m-%d",
    "%d %b %Y", "%d %B %Y",
    "%d-%b-%Y", "%d-%B-%Y",
    "%d/%b/%Y", "%d/%B/%Y",
    "%d %b %y", "%d %B %y",
]

_FORMATS_AMBIGU_DAYFIRST   = ["%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%d %m %Y", "%d %m %y",
                               "%d|%m|%Y", "%d|%m|%y"]
_FORMATS_AMBIGU_MONTHFIRST = ["%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%m-%d-%y", "%m %d %Y", "%m %d %y",
                                "%m|%d|%Y", "%m|%d|%y"]


def _angka_bagian(tgl_str: str):
    return _re.findall(r"\d+", tgl_str)


def _parse_tanggal(tgl_str: str, dayfirst: bool = True) -> tuple[datetime | None, bool]:
    tgl_str = tgl_str.strip()
    if not tgl_str or tgl_str.lower() in ("nan", "none", "-", ""):
        return None, False

    tgl_str = _ganti_bulan_id(tgl_str)

    if tgl_str.isdigit() and 10000 < int(tgl_str) < 60000:
        try:
            return datetime(1899, 12, 30) + _timedelta(days=int(tgl_str)), False
        except Exception:
            pass

    for fmt in _FORMATS_PASTI:
        try:
            return datetime.strptime(tgl_str, fmt), False
        except ValueError:
            continue

    bagian = _angka_bagian(tgl_str)
    if len(bagian) >= 2:
        angka_pertama = int(bagian[0])

        if angka_pertama > 12:
            for fmt in _FORMATS_AMBIGU_DAYFIRST:
                try:
                    return datetime.strptime(tgl_str, fmt), False
                except ValueError:
                    continue
        else:
            formats_utama = _FORMATS_AMBIGU_DAYFIRST if dayfirst else _FORMATS_AMBIGU_MONTHFIRST
            formats_alt   = _FORMATS_AMBIGU_MONTHFIRST if dayfirst else _FORMATS_AMBIGU_DAYFIRST

            for fmt in formats_utama:
                try:
                    return datetime.strptime(tgl_str, fmt), True
                except ValueError:
                    continue

            for fmt in formats_alt:
                try:
                    return datetime.strptime(tgl_str, fmt), True
                except ValueError:
                    continue

    try:
        return dateutil_parser.parse(tgl_str, dayfirst=dayfirst), False
    except Exception:
        pass

    return None, False


def tentukan_kategori_umur(usia):
    if usia is None:
        return "TIDAK VALID"
    if usia < 18:
        return "ANAK"
    if usia >= 60:
        return "LANSIA"
    return "DEWASA"


def proses_kolom_usia(df_result, col_tgl_lahir, tgl_pengecekan, dayfirst: bool = True):
    usia_col     = f"USIA_{col_tgl_lahir}"
    kategori_col = f"KATEGORI_UMUR_{col_tgl_lahir}"
    parsed_col   = f"TGL_PARSED_{col_tgl_lahir}"
    catatan_col  = f"CATATAN_PARSE_{col_tgl_lahir}"

    def _parse_row(x):
        tgl, is_ambigu = _parse_tanggal(str(x), dayfirst=dayfirst)
        if tgl is None:
            return None, None, "TIDAK DIKENALI", "Format tidak dikenali"
        if tgl > tgl_pengecekan:
            return None, None, "TIDAK DIKENALI", "Tanggal di masa depan"
        usia = (
            tgl_pengecekan.year - tgl.year
            - ((tgl_pengecekan.month, tgl_pengecekan.day) < (tgl.month, tgl.day))
        )
        if usia > 130:
            return None, None, "TIDAK DIKENALI", "Usia > 130 tahun"
        catatan = "Ambigu (dd/mm atau mm/dd?)" if is_ambigu else "OK"
        return usia, tentukan_kategori_umur(usia), tgl.strftime("%d/%m/%Y"), catatan

    hasil = df_result[col_tgl_lahir].apply(_parse_row)
    df_result[usia_col]     = hasil.apply(lambda x: x[0])
    df_result[kategori_col] = hasil.apply(lambda x: x[1] if x[1] else "TIDAK VALID")
    df_result[parsed_col]   = hasil.apply(lambda x: x[2])
    df_result[catatan_col]  = hasil.apply(lambda x: x[3])

    return df_result, usia_col, kategori_col, parsed_col, catatan_col


# ─── UI HELPERS (VALIDASI) ─────────────────────────────────────────────────────

def render_charts(df_result, col_name):
    status_col = f"STATUS_{col_name}"
    viz_series = df_result[status_col].apply(
        lambda x: "GANDA" if str(x).startswith("GANDA") else x
    )
    data_counts = viz_series.value_counts().reset_index()
    data_counts.columns = ["Status", "Jumlah"]

    total      = len(df_result)
    total_unik = (df_result[status_col] == "UNIK").sum()

    m1, m2, m3 = st.columns(3)
    m1.metric("Total Data",          total)
    m2.metric("Data Valid (UNIK)",   total_unik)
    m3.metric("Data Perlu Perbaikan", total - total_unik, delta_color="inverse")

    st.markdown("---")
    col_pie, col_bar = st.columns(2)

    with col_pie:
        fig_pie = px.pie(
            data_counts, values="Jumlah", names="Status",
            title=f"Persentase: {col_name}",
            color="Status", color_discrete_map=COLOR_MAP, hole=0.4,
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    with col_bar:
        fig_bar = px.bar(
            data_counts, x="Status", y="Jumlah",
            title=f"Jumlah Error: {col_name}",
            color="Status", color_discrete_map=COLOR_MAP, text="Jumlah",
        )
        fig_bar.update_traces(textposition="outside")
        st.plotly_chart(fig_bar, use_container_width=True)


def render_charts_kategori_umur(df_result, kategori_col, col_tgl_lahir, usia_col, parsed_col, catatan_col):
    data_counts = df_result[kategori_col].value_counts().reset_index()
    data_counts.columns = ["Kategori", "Jumlah"]

    total       = len(df_result)
    jml_anak    = (df_result[kategori_col] == "ANAK").sum()
    jml_dewasa  = (df_result[kategori_col] == "DEWASA").sum()
    jml_lansia  = (df_result[kategori_col] == "LANSIA").sum()
    jml_invalid = (df_result[kategori_col] == "TIDAK VALID").sum()

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total Data",       total)
    m2.metric("Anak",             jml_anak)
    m3.metric("Dewasa",           jml_dewasa)
    m4.metric("Lansia",           jml_lansia)
    m5.metric("Tgl Tidak Valid",  jml_invalid)

    st.markdown("---")
    col_pie, col_bar = st.columns(2)

    with col_pie:
        fig_pie = px.pie(
            data_counts, values="Jumlah", names="Kategori",
            title=f"Distribusi Kategori Umur: {col_tgl_lahir}",
            color="Kategori", color_discrete_map=COLOR_MAP_KATEGORI, hole=0.4,
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    with col_bar:
        fig_bar = px.bar(
            data_counts, x="Kategori", y="Jumlah",
            title=f"Jumlah per Kategori Umur: {col_tgl_lahir}",
            color="Kategori", color_discrete_map=COLOR_MAP_KATEGORI, text="Jumlah",
        )
        fig_bar.update_traces(textposition="outside")
        st.plotly_chart(fig_bar, use_container_width=True)

    df_valid_usia = df_result[df_result[usia_col].notna()].copy()
    if not df_valid_usia.empty:
        fig_hist = px.histogram(
            df_valid_usia, x=usia_col, color=kategori_col,
            nbins=30,
            title="Distribusi Usia",
            labels={usia_col: "Usia (Tahun)", "count": "Jumlah"},
            color_discrete_map=COLOR_MAP_KATEGORI,
        )
        fig_hist.add_vline(x=18, line_dash="dash", line_color="blue",
                           annotation_text="18 th (Dewasa)", annotation_position="top right")
        fig_hist.add_vline(x=60, line_dash="dash", line_color="red",
                           annotation_text="60 th (Lansia)", annotation_position="top right")
        st.plotly_chart(fig_hist, use_container_width=True)

    df_gagal = df_result[df_result[parsed_col] == "TIDAK DIKENALI"]
    if not df_gagal.empty:
        with st.expander(f"⚠️ {len(df_gagal)} baris tanggal tidak berhasil di-parse — klik untuk lihat detail", expanded=False):
            st.caption("Baris berikut tidak dapat dikenali formatnya. Silakan perbaiki secara manual.")
            st.dataframe(
                df_gagal[[col_tgl_lahir, parsed_col]].reset_index(drop=True),
                use_container_width=True,
            )

    df_ambigu = df_result[df_result[catatan_col] == "Ambigu (dd/mm atau mm/dd?)"]
    if not df_ambigu.empty:
        with st.expander(
            f"🔍 {len(df_ambigu)} baris berformat ambigu (angka pertama ≤ 12) — klik untuk verifikasi",
            expanded=False,
        ):
            st.caption(
                "Baris-baris ini sudah diproses sesuai pilihan interpretasi Anda, "
                "namun sebaiknya diverifikasi manual karena format dd/mm dan mm/dd tidak bisa dibedakan secara otomatis."
            )
            st.dataframe(
                df_ambigu[[col_tgl_lahir, parsed_col, catatan_col, usia_col, kategori_col]].reset_index(drop=True),
                use_container_width=True,
            )


def render_filtered_table(df_result, target_cols):
    df_display = df_result.copy()
    filter_cols = st.columns(len(target_cols))

    for idx, col_name in enumerate(target_cols):
        status_col = f"STATUS_{col_name}"
        with filter_cols[idx]:
            list_status = df_result[status_col].unique().tolist()
            pilihan = st.multiselect(
                f"Filter {status_col}:", options=list_status, default=list_status
            )
            df_display = df_display[df_display[status_col].isin(pilihan)]

    st.caption(f"Menampilkan {len(df_display)} dari total {len(df_result)} baris data.")
    st.dataframe(df_display, use_container_width=True)
    return df_display


def render_filtered_table_usia(df_result, kategori_cols):
    df_display = df_result.copy()
    filter_cols = st.columns(len(kategori_cols))

    for idx, kategori_col in enumerate(kategori_cols):
        with filter_cols[idx]:
            list_kat = df_result[kategori_col].unique().tolist()
            pilihan = st.multiselect(
                f"Filter {kategori_col}:", options=list_kat, default=list_kat
            )
            df_display = df_display[df_display[kategori_col].isin(pilihan)]

    st.caption(f"Menampilkan {len(df_display)} dari total {len(df_result)} baris data.")
    st.dataframe(df_display, use_container_width=True)
    return df_display


def buat_excel_buffer(df_result, selected_sheet):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        sheet_name = f"Cek_{selected_sheet}"[:30]
        df_result.to_excel(writer, index=False, sheet_name=sheet_name)

        wb = writer.book
        ws = writer.sheets[sheet_name]
        txt_fmt = wb.add_format({"num_format": "@"})
        for idx in range(len(df_result.columns)):
            ws.set_column(idx, idx, 25, txt_fmt)

    buffer.seek(0)
    return buffer


def bersihkan_nama_file(nama):
    for ext in (".xlsx", ".xlsm", ".xls", ".csv"):
        if nama.endswith(ext):
            return nama[: -len(ext)]
    return nama


# ─── SIDEBAR (VALIDASI) ─────────────────────────────────────────────────────────

def render_sidebar(waktu_tarik):
    with st.sidebar:
        st.info(f"⏱️ **Data Salur Terakhir Ditarik:**\n\n{waktu_tarik}")
        st.caption("Sistem mengunci memori selama 1 jam untuk mencegah blokir server Google.")
        st.divider()

        st.header("⚙️ Admin Panel")
        if st.checkbox("Lihat Log Aktivitas"):
            try:
                with open("activity_log.txt", "r") as f:
                    st.text(f.read())
            except Exception:
                st.text("Log kosong.")

        if st.button("Hapus Log"):
            try:
                with open("activity_log.txt", "w") as f:
                    pass
                st.rerun()
            except Exception:
                pass


# ─── FILE READING (VALIDASI) ───────────────────────────────────────────────────

def baca_preview_mentah(uploaded_file, selected_sheet, is_csv):
    if is_csv:
        uploaded_file.seek(0)
        try:
            return pd.read_csv(uploaded_file, header=None, nrows=10)
        except Exception:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, header=None, nrows=10, sep=";")
    else:
        return pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=None, nrows=10, engine="openpyxl")


def baca_data_penuh(uploaded_file, selected_sheet, is_csv, header_row_input):
    header_idx = header_row_input - 1
    if is_csv:
        uploaded_file.seek(0)
        try:
            return pd.read_csv(uploaded_file, header=header_idx)
        except Exception:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, header=header_idx, sep=";")
    else:
        return pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=header_idx, engine="openpyxl")


# ─── VALIDASI PAGE ─────────────────────────────────────────────────────────────

def render_validasi_page():
    st.markdown(STYLES, unsafe_allow_html=True)
    st.title("📊 Dashboard Validasi Data - Internal Antasena")
    st.info("Fitur: Atur Posisi Header, Multi-Kolom, Multi-Sheet, Auto Cleansing, Visualisasi, Auto-Format Text & Kategori Umur.")

    set_salur_2026, waktu_tarik = ambil_data_salur_gspread()
    render_sidebar(waktu_tarik)

    if "is_processed" not in st.session_state:
        st.session_state.is_processed = False

    uploaded_file = st.file_uploader("Upload file Excel/CSV", type=["xlsx", "xlsm", "xls", "csv"])
    if uploaded_file is None:
        st.write("<br><br><br>", unsafe_allow_html=True)
        return

    try:
        is_csv = uploaded_file.name.endswith(".csv")
        if is_csv:
            daftar_sheet = ["Sheet1"]
        else:
            xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
            daftar_sheet = xls.sheet_names

        st.subheader("1. Konfigurasi File")
        col_sheet, col_header_row = st.columns([2, 1])

        with col_sheet:
            if not is_csv:
                selected_sheet = st.selectbox("Pilih Sheet:", daftar_sheet)
            else:
                st.info("File CSV terdeteksi (Hanya 1 Sheet).")
                selected_sheet = daftar_sheet[0]

        df_preview_raw = baca_preview_mentah(uploaded_file, selected_sheet, is_csv)
        df_preview_raw = df_preview_raw.fillna("")

        with st.expander("🔍 Klik untuk melihat Preview Data Mentah (Cek posisi Header)", expanded=False):
            st.caption("Baris ke berapa Header tabel Anda?")
            df_preview_raw.index += 1
            st.dataframe(df_preview_raw, use_container_width=True)

        with col_header_row:
            header_row_input = st.number_input("Header Table ada di baris ke:", min_value=1, value=1)
            hapus_baris_penomoran = st.checkbox(
                "Abaikan baris nomor kolom (1, 2, 3...) "
                "Membuang 1 baris di bawah header bila terdapat urutan angka kolom",
                value=False,
                help="Otomatis membuang 1 baris tepat di bawah header jika isinya hanya urutan angka kolom.",
            )

        df = baca_data_penuh(uploaded_file, selected_sheet, is_csv, header_row_input)
        df.dropna(how="all", inplace=True)

        if hapus_baris_penomoran and not df.empty:
            df = df.iloc[1:].reset_index(drop=True)

        df = df.astype(str)
        for col in df.columns:
            df[col] = df[col].replace("nan", "").str.replace(r"\.0$", "", regex=True)

        st.divider()
        st.subheader("2. Pilih Kolom Data")
        cols = df.columns.tolist()

        if not cols:
            st.error("⚠️ Header tidak ditemukan.")
            return

        col_left, col_right = st.columns([3, 1])
        with col_left:
            target_cols = st.multiselect(
                "Pilih Kolom yang akan dicek (NIK/NKK):",
                cols,
                placeholder="Pilih kolom NIK, KK, dll...",
            )
        with col_right:
            use_auto_clean = st.checkbox(
                "Aktifkan Auto-Cleaning",
                value=False,
                help="Otomatis menghapus spasi, titik, strip, dan huruf.",
            )

        st.divider()
        st.subheader("3. Cek Kategori Umur (Opsional)")
        st.caption(
            "💡 Fitur ini **dapat dipilih bila terdapat kolom tanggal lahir** dari PM tersebut. "
            "Jika file tidak memiliki isian tanggal lahir, abaikan bagian ini — "
            "proses validasi NIK/NKK tetap berjalan seperti biasa."
        )

        aktifkan_cek_umur = st.checkbox(
            "Aktifkan Pengecekan Kategori Umur (Anak / Dewasa / Lansia)",
            value=False,
            help="Hanya centang jika file Anda memiliki kolom tanggal lahir.",
        )

        cols_tgl_lahir_dipilih = []
        tgl_pengecekan = datetime.now(ZoneInfo("Asia/Jakarta")).replace(tzinfo=None)
        dayfirst = True

        if aktifkan_cek_umur:
            col_umur_left, col_umur_right = st.columns([3, 1])
            with col_umur_left:
                cols_tgl_lahir_dipilih = st.multiselect(
                    "Pilih Kolom Tanggal Lahir:",
                    cols,
                    placeholder="Pilih kolom yang berisi tanggal lahir...",
                    help="Format tanggal yang didukung: dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd, serial Excel, dll.",
                )
            with col_umur_right:
                tgl_pengecekan_input = st.date_input(
                    "Tanggal Pengecekan:",
                    value=datetime.now(ZoneInfo("Asia/Jakarta")).date(),
                    format="DD/MM/YYYY",
                    help="Usia akan dihitung berdasarkan tanggal ini.",
                )
                tgl_pengecekan = datetime.combine(tgl_pengecekan_input, datetime.min.time())

            if not cols_tgl_lahir_dipilih:
                st.info("ℹ️ Pilih kolom tanggal lahir di atas untuk mengaktifkan analisa kategori umur.")
            else:
                st.warning(
                    "⚠️ **Perhatian Format Ambigu**\n\n"
                    "Tanggal seperti `03/05/1990` bisa berarti **3 Mei** (dd/mm) atau **5 Maret** (mm/dd). "
                    "Pilih interpretasi default di bawah untuk kasus seperti ini."
                )
                interpretasi_ambigu = st.radio(
                    "Jika format tanggal ambigu, interpretasikan angka pertama sebagai:",
                    options=["Hari (dd/mm/yyyy) — default Indonesia", "Bulan (mm/dd/yyyy) — gaya Amerika"],
                    index=0,
                    horizontal=True,
                )
                dayfirst = interpretasi_ambigu.startswith("Hari")

                st.info(
                    f"📌 **Aturan Kategorisasi:**\n"
                    f"- **ANAK**: Usia < 18 tahun\n"
                    f"- **DEWASA**: 18 \u2264 Usia < 60 tahun\n"
                    f"- **LANSIA**: Usia \u2264 60 tahun\n\n"
                    f"Tanggal pengecekan: **{tgl_pengecekan_input.strftime('%d/%m/%Y')}** | "
                    f"Interpretasi ambigu: **{'dd/mm' if dayfirst else 'mm/dd'}**"
                )

        if st.button("🚀 Proses & Analisa Data"):
            if not target_cols:
                st.warning("⚠️ Silakan pilih minimal 1 kolom NIK/NKK untuk diproses.")
            else:
                with st.spinner("Memproses data..."):
                    df_result = df.copy()
                    log_data_all = {}

                    for col_name in target_cols:
                        df_result = proses_kolom(df_result, col_name, use_auto_clean, set_salur_2026)
                        log_data_all[col_name] = df_result[f"STATUS_{col_name}"].value_counts().to_dict()

                    hasil_usia = {}
                    if aktifkan_cek_umur and cols_tgl_lahir_dipilih:
                        for col_tgl in cols_tgl_lahir_dipilih:
                            df_result, usia_col, kategori_col, parsed_col, catatan_col = proses_kolom_usia(
                                df_result, col_tgl, tgl_pengecekan, dayfirst=dayfirst
                            )
                            hasil_usia[col_tgl] = (usia_col, kategori_col, parsed_col, catatan_col)

                    catat_log(uploaded_file.name, selected_sheet, log_data_all)

                    st.session_state.df_result = df_result
                    st.session_state.target_cols_saved = target_cols
                    st.session_state.hasil_usia = hasil_usia
                    st.session_state.is_processed = True

        if (
            st.session_state.get("is_processed")
            and st.session_state.get("target_cols_saved") == target_cols
        ):
            df_result = st.session_state.df_result
            hasil_usia = st.session_state.get("hasil_usia", {})

            if target_cols:
                st.divider()
                st.subheader("📊 Hasil Analisa Visual NIK/NKK")

                tabs = st.tabs([f"Analisa: {c}" for c in target_cols])
                for i, col_name in enumerate(target_cols):
                    with tabs[i]:
                        render_charts(df_result, col_name)

                st.divider()
                st.subheader("📋 Tabel Data NIK/NKK")
                render_filtered_table(df_result, target_cols)

            if hasil_usia:
                st.divider()
                st.subheader("👥 Hasil Analisa Kategori Umur")

                tab_labels = [f"Umur: {col}" for col in hasil_usia.keys()]
                tabs_umur = st.tabs(tab_labels)

                for i, (col_tgl, (usia_col, kategori_col, parsed_col, catatan_col)) in enumerate(hasil_usia.items()):
                    with tabs_umur[i]:
                        render_charts_kategori_umur(df_result, kategori_col, col_tgl, usia_col, parsed_col, catatan_col)

                st.divider()
                st.subheader("📋 Tabel Data Kategori Umur")
                kategori_cols = [v[1] for v in hasil_usia.values()]
                render_filtered_table_usia(df_result, kategori_cols)

            st.divider()
            buffer = buat_excel_buffer(df_result, selected_sheet)
            clean_name = bersihkan_nama_file(uploaded_file.name)
            st.download_button(
                label="📥 Download Hasil Seluruhnya (Excel)",
                data=buffer,
                file_name=f"Result_{clean_name}.xlsx",
                mime="application/vnd.ms-excel",
            )

    except Exception as e:
        st.error(f"Terjadi kesalahan: {e}")

    st.write("<br><br><br>", unsafe_allow_html=True)


# ─── SPLIT WORKBOOK PAGE ───────────────────────────────────────────────────────

def _get_help_text(section: str) -> str:
    help_content = {
        "upload": "Upload file Excel (.xlsx, .xlsm, .xls). File akan diproses di memory server dan tidak disimpan secara permanen.",
        "header_row": "Baris yang berisi nama kolom header. Default: 1 (baris pertama). Ubah jika header tidak di baris pertama.",
        "kolom_split": "Pilih kolom yang nilainya akan digunakan untuk memecah file. Setiap unique value di kolom ini akan menjadi 1 file output terpisah.",
        "preview_stats": "Menampilkan statistik dasar: jumlah baris data, jumlah unique value di kolom split, dan estimasi jumlah file yang akan dihasilkan.",
        "auto_clean": "Gabungkan nilai yang serupa menjadi satu. Contoh: 'kab. boyolali', 'kabupaten boyolali', 'KABUPATEN BOYOLALI' akan digabung menjadi satu nilai. Menggunakan fuzzy matching dengan threshold 80%.",
        "text_format": "Aktifkan untuk menjaga format teks di kolom tertentu agar tidak terkonversi menjadi angka oleh Excel. Contoh: NIK '0012345678901234' tidak berubah jadi '1.2345E+15'.",
        "select_columns": "Pilih kolom yang perlu diformat sebagai teks. Sistem akan otomatis mendeteksi kolom yang mengandung keyword seperti: NIK, KK, NO. HP, TANGGAL, SK, BAST.",
        "select_all": "Centang untuk memilih semua kolom. Batalkan centang untuk membatalkan semua pilihan.",
        "process": "Memulai proses pemisahan file. File output akan dibundle dalam 1 file ZIP untuk download.",
        "cancel": "Membatalkan proses. Semua file yang sudah dibuat akan dihapus.",
    }
    return help_content.get(section, "")


def _auto_detect_text_columns(columns: list) -> set:
    detected = set()
    for col in columns:
        col_upper = str(col).upper()
        for keyword in TEXT_COLUMNS_KEYWORDS:
            if keyword.upper() in col_upper:
                detected.add(col)
                break
    return detected


def _normalize_for_comparison(val: str) -> str:
    """Normalize value for fuzzy comparison."""
    val = str(val).lower().strip()
    val = val.replace(".", "").replace(",", "")
    val = val.replace("  ", " ")
    return val


def _get_similarity(s1: str, s2: str) -> float:
    """Calculate similarity ratio between two strings."""
    from difflib import SequenceMatcher
    return SequenceMatcher(None, s1, s2).ratio()


def _fuzzy_group_values(unique_values: list, frequency_map: dict, threshold: float = 0.80) -> dict:
    """
    Group similar values using fuzzy matching.
    Returns dict: {winner_value: [list of member values]}
    """
    if not unique_values:
        return {}
    
    values_to_check = list(unique_values)
    clusters = {}
    used = set()
    
    for i, val in enumerate(values_to_check):
        if val in used:
            continue
        
        cluster = [val]
        used.add(val)
        val_normalized = _normalize_for_comparison(val)
        
        for j, other in enumerate(values_to_check):
            if other in used or i == j:
                continue
            
            other_normalized = _normalize_for_comparison(other)
            
            similarity = _get_similarity(val_normalized, other_normalized)
            
            if similarity >= threshold:
                cluster.append(other)
                used.add(other)
        
        winner = max(cluster, key=lambda x: frequency_map.get(x, 0))
        clusters[winner] = cluster
    
    return clusters


def _apply_cleaning_to_df(df: pd.DataFrame, col: str, clusters: dict) -> pd.DataFrame:
    """Apply cleaning by replacing values with their winner."""
    df = df.copy()
    value_mapping = {}
    for winner, members in clusters.items():
        for member in members:
            value_mapping[member] = winner
    df[col] = df[col].replace(value_mapping)
    return df


def _enforce_text_format_in_memory(excel_bytes: bytes, sheet_name: str, selected_columns: set) -> bytes:
    from io import BytesIO
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb[sheet_name]
    
    col_indices = {}
    for col_cell in ws[1]:
        if col_cell.value in selected_columns:
            col_indices[col_cell.column] = col_cell.value
    
    if col_indices:
        for col_idx in col_indices:
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        cell.value = str(cell.value).strip()
                        cell.number_format = "@"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def render_split_page():
    st.title("📤 Split Workbook - Internal Antasena")
    st.caption("Pecah file Excel berdasarkan kolom tertentu menjadi beberapa file.")

    if "split_state" not in st.session_state:
        st.session_state.split_state = {
            "columns_loaded": False,
            "df_preview": None,
            "all_columns": [],
            "processing": False,
            "cancel_requested": False,
            "progress": 0,
            "files_created": [],
            "start_time": None,
        }

    uploaded_file = st.file_uploader(
        "📁 Upload File Excel",
        type=["xlsx", "xlsm", "xls"],
        help=_get_help_text("upload")
    )

    if not uploaded_file:
        st.write("<br><br>", unsafe_allow_html=True)
        return

    try:
        is_csv = uploaded_file.name.endswith(".csv")
        
        st.subheader("1. Konfigurasi File")
        col_file, col_header_row = st.columns([3, 1])
        
        with col_file:
            if not is_csv:
                xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
                daftar_sheet = xls.sheet_names
                selected_sheet = st.selectbox("📋 Sheet:", daftar_sheet)
            else:
                selected_sheet = "Sheet1"
                st.info("File CSV terdeteksi (Hanya 1 Sheet).")
        
        df_preview_raw = baca_preview_mentah(uploaded_file, selected_sheet, is_csv)
        df_preview_raw = df_preview_raw.fillna("")
        
        with st.expander("🔍 Klik untuk melihat Preview Data Mentah (Cek posisi Header)", expanded=False):
            st.caption("Baris ke berapa Header tabel Anda?")
            df_preview_raw.index += 1
            st.dataframe(df_preview_raw, use_container_width=True)
        
        with col_header_row:
            header_row_input = st.number_input(
                "Header Table ada di baris ke:",
                min_value=1,
                value=1,
                help=_get_help_text("header_row")
            )
            hapus_baris_penomoran = st.checkbox(
                "Abaikan baris nomor kolom (1, 2, 3...) "
                "Membuang 1 baris di bawah header bila terdapat urutan angka kolom",
                value=False,
                help="Otomatis membuang 1 baris tepat di bawah header jika isinya hanya urutan angka kolom.",
            )
        
        df_full = baca_data_penuh(uploaded_file, selected_sheet, is_csv, header_row_input)
        df_full.dropna(how="all", inplace=True)
        
        if hapus_baris_penomoran and not df_full.empty:
            df_full = df_full.iloc[1:].reset_index(drop=True)
        
        df_full = df_full.astype(str)
        for col in df_full.columns:
            df_full[col] = df_full[col].replace("nan", "").str.replace(r"\.0$", "", regex=True)
        
        st.divider()
        
        st.subheader("2. Pilih Kolom Split")
        cols = df_full.columns.tolist()
        
        if not cols:
            st.error("⚠️ Header tidak ditemukan.")
            return
        
        col_split = st.selectbox(
            "📌 Kolom Split:",
            cols,
            help=_get_help_text("kolom_split")
        )
        
        target_split_col = col_split

        if target_split_col and target_split_col in df_full.columns:
            st.divider()
            st.subheader("📊 Preview Statistik")
            
            df_stats = df_full
            
            unique_vals = [v for v in df_stats[target_split_col].unique() if str(v).strip() not in ("", "nan", "None")]
            freq_map = df_stats[target_split_col].value_counts().to_dict()
            
            col_stat1, col_stat2, col_stat3 = st.columns(3)
            col_stat1.metric("Total Baris Data", len(df_stats))
            col_stat2.metric("Unique Nilai Split", len(unique_vals))
            col_stat3.metric("Estimasi File Output", len(unique_vals))
            
            if len(unique_vals) > 100:
                st.warning(f"⚠️ Perhatian: Akan ada {len(unique_vals)} file output. Proses mungkin memerlukan waktu lama.")
            
            st.info(_get_help_text("preview_stats"))
            
            enable_auto_clean = st.checkbox(
                "☑ Aktifkan Auto Cleaning (Fuzzy Match)",
                value=False,
                help="Gabungkan nilai yang serupa (misal: 'kab. boyolali' dan 'kabupaten boyolali') menjadi satu"
            )
            
            clusters = {}
            cleaned_unique_count = len(unique_vals)
            
            if enable_auto_clean:
                clusters = _fuzzy_group_values(unique_vals, freq_map, threshold=0.80)
                cleaned_unique_count = len(clusters)
            
            if enable_auto_clean and clusters:
                st.divider()
                st.subheader(f"📋 Preview Auto Cleaning: {target_split_col}")
                
                cluster_list = list(clusters.items())
                cluster_list.sort(key=lambda x: len(x[1]), reverse=True)
                
                col_stat_clean1, col_stat_clean2, col_stat_clean3 = st.columns(3)
                col_stat_clean1.metric("Klaster Ditemukan", len(clusters))
                col_stat_clean2.metric("Estimasi Penggabungan", len(unique_vals) - cleaned_unique_count)
                col_stat_clean3.metric("Estimasi File Output", cleaned_unique_count)
                
                st.caption("💡 Suggestion = nilai yang paling sering muncul di data")
                
                PREVIEW_LIMIT = 15
                show_all = st.checkbox("Tampilkan semua klaster", value=False)
                
                user_winner_picks = {}
                
                preview_clusters = cluster_list[:PREVIEW_LIMIT] if not show_all else cluster_list
                remaining_clusters = cluster_list[PREVIEW_LIMIT:] if not show_all else []
                
                for winner, members in preview_clusters:
                    if len(members) > 1:
                        with st.container():
                            st.markdown(f"**Cluster ({len(members)} nilai):**")
                            member_display = ", ".join([f"`{m}`" for m in members])
                            st.caption(member_display)
                            
                            freq_sorted = sorted(members, key=lambda x: freq_map.get(x, 0), reverse=True)
                            options = freq_sorted
                            
                            default_idx = 0
                            selected = st.selectbox(
                                f"Pilih winner:",
                                options=options,
                                index=default_idx,
                                key=f"winner_{winner}"
                            )
                            user_winner_picks[winner] = selected
                            
                            with st.expander(" atau ketik manual...", expanded=False):
                                manual_input = st.text_input(
                                    "Ketik winner manual:",
                                    value="",
                                    key=f"manual_{winner}"
                                )
                                if manual_input.strip():
                                    user_winner_picks[winner] = manual_input.strip()
                            
                            st.markdown("---")
                
                if remaining_clusters:
                    with st.expander(f"▶ Lihat {len(remaining_clusters)} klaster lainnya"):
                        for winner, members in remaining_clusters:
                            if len(members) > 1:
                                with st.container():
                                    st.markdown(f"**Cluster ({len(members)} nilai):**")
                                    member_display = ", ".join([f"`{m}`" for m in members])
                                    st.caption(member_display)
                                    
                                    freq_sorted = sorted(members, key=lambda x: freq_map.get(x, 0), reverse=True)
                                    options = freq_sorted
                                    
                                    default_idx = 0
                                    selected = st.selectbox(
                                        f"Pilih winner:",
                                        options=options,
                                        index=default_idx,
                                        key=f"winner_{winner}_remaining"
                                    )
                                    user_winner_picks[winner] = selected
                                    
                                    with st.expander(" atau ketik manual...", expanded=False):
                                        manual_input = st.text_input(
                                            "Ketik winner manual:",
                                            value="",
                                            key=f"manual_{winner}_remaining"
                                        )
                                        if manual_input.strip():
                                            user_winner_picks[winner] = manual_input.strip()
                                    
                                    st.markdown("---")
                
                final_clusters = {}
                for winner, members in clusters.items():
                    picked_winner = user_winner_picks.get(winner, winner)
                    if picked_winner not in final_clusters:
                        final_clusters[picked_winner] = []
                    for m in members:
                        if m != picked_winner:
                            final_clusters[picked_winner].append(m)
                
                cleaned_unique_count = len(final_clusters)

            st.divider()
            st.subheader("⚙️ Pengaturan Format")
            
            enable_text_format = st.checkbox(
                "Aktifkan format teks untuk kolom sensitif",
                value=False,
                help=_get_help_text("text_format")
            )
            
            all_columns = cols
            
            if enable_text_format:
                checked_columns = st.multiselect(
                    "Pilih kolom yang perlu diformat teks (agar tidak terkonversi):",
                    options=all_columns,
                    default=[],
                    help=_get_help_text("select_columns")
                )
            else:
                checked_columns = []

            st.divider()

            col_process, col_cancel = st.columns([1, 1])
            
            with col_process:
                process_disabled = st.session_state.split_state["processing"]
                btn_proses = st.button(
                    "🚀 JALANKAN PROSES",
                    use_container_width=True,
                    disabled=process_disabled
                )
            
            with col_cancel:
                if st.session_state.split_state["processing"]:
                    if st.button("⏹️ BATALKAN", use_container_width=True):
                        st.session_state.split_state["cancel_requested"] = True
                        st.rerun()

            if btn_proses and target_split_col:
                st.session_state.split_state["processing"] = True
                st.session_state.split_state["cancel_requested"] = False
                st.session_state.split_state["progress"] = 0
                st.session_state.split_state["files_created"] = []
                st.session_state.split_state["start_time"] = time.time()
                
                progress_bar = st.progress(0)
                progress_text = st.empty()
                status_text = st.empty()
                
                try:
                    df_split_data = df_full
                    split_column = target_split_col
                    
                    if enable_auto_clean and final_clusters:
                        df_split_data = _apply_cleaning_to_df(df_split_data, split_column, final_clusters)
                    
                    df_split_data = df_split_data.fillna("")
                    
                    unique_vals = [v for v in df_split_data[split_column].unique() if str(v).strip() not in ("", "nan", "None")]
                    total_files = len(unique_vals)
                    
                    zip_buffer = io.BytesIO()
                    
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for i, val in enumerate(unique_vals):
                            if st.session_state.split_state["cancel_requested"]:
                                status_text.warning("⚠️ Proses dibatalkan. Semua file yang sudah dibuat akan dihapus.")
                                break
                            
                            df_subset = df_split_data[df_split_data[split_column] == val].reset_index(drop=True)
                            
                            safe_name = str(val).strip()
                            for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
                                safe_name = safe_name.replace(char, '-')
                            
                            filename = f"{safe_name}.xlsx"
                            
                            temp_buffer = io.BytesIO()
                            with pd.ExcelWriter(temp_buffer, engine='openpyxl') as writer:
                                df_subset.to_excel(writer, index=False, sheet_name=selected_sheet)
                            temp_buffer.seek(0)
                            
                            if checked_columns:
                                excel_bytes = _enforce_text_format_in_memory(
                                    temp_buffer.getvalue(),
                                    selected_sheet,
                                    set(checked_columns)
                                )
                                zf.writestr(filename, excel_bytes)
                            else:
                                zf.writestr(filename, temp_buffer.getvalue())
                            
                            st.session_state.split_state["files_created"].append(filename)
                            
                            progress = int((i + 1) / total_files * 100)
                            st.session_state.split_state["progress"] = progress
                            
                            progress_bar.progress(progress)
                            progress_text.text(f"{i + 1}/{total_files} files ({progress}%)")
                    
                    if not st.session_state.split_state["cancel_requested"]:
                        elapsed = time.time() - st.session_state.split_state["start_time"]
                        
                        zip_buffer.seek(0)
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        
                        st.divider()
                        st.success(f"✅ Selesai! {len(st.session_state.split_state['files_created'])} file berhasil dibuat ({elapsed:.1f} detik)")
                        
                        st.download_button(
                            label="📥 Download Hasil Split (ZIP)",
                            data=zip_buffer.getvalue(),
                            file_name=f"split_result_{timestamp}.zip",
                            mime="application/zip",
                            use_container_width=True
                        )
                    else:
                        zip_buffer.close()
                        
                except Exception as e:
                    st.error(f"Terjadi kesalahan: {e}")
                
                finally:
                    st.session_state.split_state["processing"] = False
            
            if st.session_state.split_state["processing"]:
                progress_bar = st.progress(st.session_state.split_state["progress"])
                progress_text = st.text(f"{st.session_state.split_state['progress']}%")
                
                if st.session_state.split_state["files_created"]:
                    status_text.text(f"Sedang memproses: {len(st.session_state.split_state['files_created'])} files sudah dibuat...")

    except Exception as e:
        st.error(f"Terjadi kesalahan: {e}")

    st.write("<br><br>", unsafe_allow_html=True)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    st.markdown(STYLES, unsafe_allow_html=True)
    
    st.sidebar.title("📋 Menu Utama")
    
    menu_options = ["Validasi Data", "Split Workbook"]
    selected_menu = st.sidebar.radio("Pilih Menu:", menu_options, index=0)
    
    if selected_menu == "Validasi Data":
        render_validasi_page()
    elif selected_menu == "Split Workbook":
        render_split_page()

    st.markdown(STYLES, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
